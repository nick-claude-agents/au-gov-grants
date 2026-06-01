"""
GrantConnect grant-opportunity scraper
Scrapes all current grant opportunities from www.grants.gov.au (GrantConnect),
captures contact email addresses, maintains a persistent Excel registry of known
emails, writes an HTML dashboard, and emails a daily report to
nick.chapman@parbery.com.au highlighting any new opportunities and addresses.

Sister script to tenders_scraper.py (AusTender). Same architecture, adapted to
GrantConnect's /Go/List + /Go/Show?GoUuid=... structure.

Config: set GMAIL_APP_PASSWORD environment variable, or edit CONFIG below.
"""

import asyncio
import html as html_mod
import json
import os
import re
import smtplib
import logging
import sys
from datetime import date as date_type
from zoneinfo import ZoneInfo

AEST = ZoneInfo("Australia/Sydney")
from datetime import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from playwright.async_api import async_playwright, Page

# ── Configuration ──────────────────────────────────────────────────────────────
CONFIG = {
    "from_email": "nick.claude.agents@gmail.com",
    "to_email": "nick.chapman@parbery.com.au",
    "gmail_app_password": "",  # Set here OR via GMAIL_APP_PASSWORD env var
    "base_url": "https://www.grants.gov.au",
    "listing_url": "https://www.grants.gov.au/Go/List",
    "concurrency": 8,
    "page_load_timeout": 40000,
    "log_file": str(Path(__file__).parent / "grants_scraper.log"),
    "registry_file": str(Path(__file__).parent / "grants_email_registry.xlsx"),
    "known_gos_file": str(Path(__file__).parent / "grants_known_gos.json"),
    "pages_url": "https://nick-claude-agents.github.io/au-gov-grants/",
    # In GitHub Actions (the au-gov-grants repo) write index.html for Pages.
    # Locally, write grants_latest.html so we don't clobber the corporate-plans
    # dashboard's index.html that shares this Dropbox folder.
    "html_file": str(Path(__file__).parent / ("index.html" if os.environ.get("GITHUB_ACTIONS") else "grants_latest.html")),
}

EMAIL_RE = re.compile(r'\b[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}\b')
EXCLUDE_DOMAINS = {"grants.gov.au", "example.com"}
# GrantConnect's own support address appears as boilerplate on every page.
EXCLUDE_EMAILS = {"grantconnect@finance.gov.au"}

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s",
    handlers=[
        logging.FileHandler(CONFIG["log_file"], encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger(__name__)


# ── Known GOs tracking ────────────────────────────────────────────────────────

def load_known_go_urls() -> set[str]:
    path = Path(CONFIG["known_gos_file"])
    if not path.exists():
        return set()
    with open(path, encoding="utf-8") as f:
        return set(json.load(f))


def save_known_go_urls(urls: list[str]) -> None:
    with open(CONFIG["known_gos_file"], "w", encoding="utf-8") as f:
        json.dump(sorted(urls), f)


# ── Excel registry ────────────────────────────────────────────────────────────

HEADERS = ["Email Address", "Agency", "GO ID", "Date Added"]

FILL_HEADER  = PatternFill("solid", fgColor="1C1C1C")
FILL_NEW     = PatternFill("solid", fgColor="FFFACD")   # lemon yellow — new this run
FILL_KNOWN   = PatternFill("solid", fgColor="FFFFFF")
FONT_HEADER  = Font(color="FFFFFF", bold=True)
FONT_NEW     = Font(bold=True, color="B8860B")           # dark gold
THIN_BORDER  = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
COL_WIDTHS   = [34, 50, 14, 14]


def _load_registry() -> dict[str, dict]:
    """Return {email_lower: row_dict} for all emails in the registry file."""
    path = Path(CONFIG["registry_file"])
    if not path.exists():
        return {}
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    known: dict[str, dict] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            known[str(row[0]).lower()] = {
                "email": row[0],
                "agency": row[1],
                "go_id": row[2],
                "date_added": row[3],
            }
    return known


def update_registry(results: list[dict]) -> tuple[set[str], set[str]]:
    """
    Update the Excel registry with today's scrape results.
    Returns (new_emails, known_emails) as sets of lowercase addresses.
    """
    existing = _load_registry()
    today = datetime.now(AEST).strftime("%Y-%m-%d")
    new_emails: set[str] = set()
    all_seen: set[str] = set()

    for result in results:
        for email in result["emails"]:
            el = email.lower()
            all_seen.add(el)
            if el not in existing:
                new_emails.add(el)
                existing[el] = {
                    "email": email,
                    "agency": result["agency"],
                    "go_id": result["go_id"],
                    "date_added": today,
                }

    # Rebuild workbook sorted by date_added desc, then email
    all_rows = sorted(
        existing.values(),
        key=lambda r: (r.get("date_added", ""), r.get("email", "")),
        reverse=True,
    )

    wb = openpyxl.Workbook()
    ws = wb.active

    # Header row
    for col, (header, width) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"

    # Data rows
    for row_idx, r in enumerate(all_rows, 2):
        is_new = r["email"].lower() in new_emails
        values = [r.get("email", ""), r.get("agency", ""), r.get("go_id", ""), r.get("date_added", "")]
        fill = FILL_NEW if is_new else FILL_KNOWN
        font = FONT_NEW if is_new else Font()
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.fill = fill
            cell.font = font
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row_idx].height = 16

    ws.title = f"Registry ({len(all_rows)} emails)"

    wb.save(CONFIG["registry_file"])
    log.info(f"Registry updated: {len(all_rows)} total, {len(new_emails)} new -> {CONFIG['registry_file']}")
    return new_emails, all_seen - new_emails


# ── Step 1: Collect all GO opportunities (url + title) ────────────────────────

async def collect_all_go_opportunities(context) -> list[dict]:
    """
    Paginate through the Grant Opportunity listing and return all unique
    opportunities as dicts: {'url': ..., 'title': ...}.
    The descriptive title is only on the listing card (the detail page <h1>
    just says 'Current Grant Opportunity View - GOxxxx'), so we capture it here.
    """
    page = await context.new_page()
    all_opps: list[dict] = []
    seen: set[str] = set()

    try:
        log.info("Loading Grant Opportunity listing page 1...")
        await page.goto(CONFIG["listing_url"], wait_until="networkidle",
                        timeout=CONFIG["page_load_timeout"])

        pag_numbers = await page.eval_on_selector_all(
            ".pagination a, [class*='pag'] a",
            "els => els.map(e => e.textContent.trim()).filter(t => /^\\d+$/.test(t))"
        )
        total_pages = max((int(n) for n in pag_numbers), default=1)
        log.info(f"Found {total_pages} listing pages")

        for pg in range(1, total_pages + 1):
            if pg > 1:
                log.info(f"Loading listing page {pg}/{total_pages}...")
                await page.goto(
                    f"{CONFIG['listing_url']}?page={pg}",
                    wait_until="networkidle",
                    timeout=CONFIG["page_load_timeout"],
                )

            cards = await page.evaluate("""(base) => {
                const out = [];
                const links = [...document.querySelectorAll("a[href*='/Go/Show']")];
                const seen = new Set();
                for (const a of links) {
                    const href = a.getAttribute('href');
                    const m = href && href.match(/GoUuid=([0-9a-fA-F-]+)/);
                    if (!m) continue;
                    const uuid = m[1];
                    if (seen.has(uuid)) continue;
                    seen.add(uuid);
                    const card = a.closest('article') || a.closest('.row');
                    let title = '';
                    if (card) {
                        const h = card.querySelector('h1,h2,h3,h4,h5');
                        title = h ? h.textContent.trim()
                                  : (card.innerText.split('\\n')[0] || '').trim();
                    }
                    const url = new URL(href, base).href;
                    out.push({uuid, url, title});
                }
                return out;
            }""", CONFIG["base_url"])

            new_on_page = 0
            for c in cards:
                if c["url"] not in seen:
                    seen.add(c["url"])
                    all_opps.append({"url": c["url"], "title": c["title"]})
                    new_on_page += 1
            log.info(f"  Page {pg}: {new_on_page} new GO opportunities (total: {len(all_opps)})")

    finally:
        await page.close()

    log.info(f"Collected {len(all_opps)} grant opportunities")
    return all_opps


# ── Step 2: Extract emails and metadata from a detail page ────────────────────

def _parse_close_date(raw: str) -> dict:
    """
    Parse a GrantConnect close date string like '1-Jun-2026 3:00 pm (ACT Local Time)'.
    Returns {'display': '01 Jun 2026', 'sort': '2026-06-01', 'raw': raw}.
    """
    m = re.match(r'(\d{1,2})-([A-Za-z]{3})-(\d{4})', raw)
    if m:
        day, mon, year = m.group(1), m.group(2), m.group(3)
        months = {"Jan":"01","Feb":"02","Mar":"03","Apr":"04","May":"05","Jun":"06",
                  "Jul":"07","Aug":"08","Sep":"09","Oct":"10","Nov":"11","Dec":"12"}
        mm = months.get(mon.capitalize(), "00")
        return {
            "display": f"{int(day):02d} {mon} {year}",
            "sort": f"{year}-{mm}-{int(day):02d}",
            "raw": raw,
        }
    return {"display": raw, "sort": "9999-99-99", "raw": raw}


async def extract_from_detail(context, opp: dict) -> dict:
    """Load one GO detail page and return email addresses and metadata found."""
    url = opp["url"]
    page = await context.new_page()
    emails: set[str] = set()
    go_id = ""
    agency = ""
    title = opp.get("title", "")
    category = ""
    close_date = {"display": "", "sort": "9999-99-99", "raw": ""}

    try:
        await page.goto(url, wait_until="networkidle", timeout=CONFIG["page_load_timeout"])
        content = await page.content()

        found = EMAIL_RE.findall(content)
        emails = {
            e.lower() for e in found
            if e.split("@")[-1].lower() not in EXCLUDE_DOMAINS
            and e.lower() not in EXCLUDE_EMAILS
        }

        # Metadata lives in <div class="list-desc"><span>Label:</span>
        #   <div class="list-desc-inner">Value</div></div> blocks.
        meta = await page.evaluate("""() => {
            const fields = {};
            document.querySelectorAll('.list-desc').forEach(d => {
                const label = d.querySelector('span');
                const val = d.querySelector('.list-desc-inner');
                if (label && val) {
                    const key = label.textContent.trim().replace(/:\\s*$/, '');
                    fields[key] = val.textContent.replace(/\\s+/g, ' ').trim();
                }
            });
            return fields;
        }""")
        go_id = meta.get("GO ID", "")
        agency = meta.get("Agency", "")
        category = meta.get("Primary Category", "")
        close_date = _parse_close_date(meta.get("Close Date & Time", "").strip())

    except Exception as e:
        log.warning(f"  {url}: failed to load ({e})")
    finally:
        await page.close()

    if emails:
        log.info(f"  {go_id or url}: {emails}")

    return {
        "url": url,
        "go_id": go_id,
        "agency": agency,
        "title": title,
        "category": category,
        "close_date": close_date,
        "emails": sorted(emails),
    }


# ── Step 3: Scrape all detail pages concurrently ──────────────────────────────

async def scrape_all_details(context, opps: list[dict]) -> list[dict]:
    """Scrape all GO detail pages with bounded concurrency."""
    sem = asyncio.Semaphore(CONFIG["concurrency"])
    results = []

    async def bounded(opp):
        async with sem:
            return await extract_from_detail(context, opp)

    tasks = [asyncio.create_task(bounded(opp)) for opp in opps]
    for i, coro in enumerate(asyncio.as_completed(tasks), 1):
        result = await coro
        results.append(result)
        if i % 10 == 0:
            log.info(f"Progress: {i}/{len(opps)} detail pages done")

    return results


# ── Step 4: Send email ────────────────────────────────────────────────────────

def send_email(results: list[dict], new_emails: set[str], new_go_urls: set[str], password: str):
    """Compose and send the results email, highlighting new opportunities & addresses."""
    run_date = datetime.now(AEST).strftime("%d %B %Y")
    with_emails = [r for r in results if r["emails"]]
    all_emails_set = {e for r in with_emails for e in r["emails"]}

    # Build table rows — ALL opportunities, sorted oldest closing date first
    rows_html = ""
    td = "padding:4px 8px;border:1px solid #d0e0dc"
    for r in sorted(results, key=lambda x: x["close_date"]["sort"]):
        is_new_go = r["url"] in new_go_urls
        is_new_email = any(e.lower() in new_emails for e in r["emails"])
        row_bg = "#fff3f4" if (is_new_go or is_new_email) else ""
        row_style = f"background:{row_bg}" if row_bg else ""
        new_label = (
            "<span style='background:#ffadb5;color:#8b0030;font-size:11px;"
            "font-weight:bold;padding:1px 5px;border-radius:3px;margin-left:6px'>NEW</span>"
        ) if is_new_go else ""

        if r["emails"]:
            email_cells = []
            for email in r["emails"]:
                if email.lower() in new_emails:
                    email_cells.append(
                        f"<span style='background:#ffadb5;color:#8b0030;font-weight:bold;"
                        f"padding:1px 4px;border-radius:3px' title='New contact'>★ {email}</span>"
                    )
                else:
                    email_cells.append(f"<a href='mailto:{email}'>{email}</a>")
            emails_html = ", ".join(email_cells)
        else:
            emails_html = "<span style='color:#bbb'>—</span>"

        rows_html += (
            f"<tr style='{row_style}'>"
            f"<td style='{td}'><a href='{r['url']}'>{r['go_id']}</a>{new_label}</td>"
            f"<td style='{td}'>{r['agency']}</td>"
            f"<td style='{td}'>{r['title']}</td>"
            f"<td style='{td};white-space:nowrap'>{r['close_date']['display']}</td>"
            f"<td style='{td}'>{emails_html}</td>"
            f"</tr>\n"
        )

    new_go_badge = ""
    if new_go_urls:
        new_go_badge = (
            f"&nbsp;|&nbsp;<strong style='color:#8b0030'>★ {len(new_go_urls)} new "
            f"opportunit{'ies' if len(new_go_urls) != 1 else 'y'} today</strong>"
        )
    new_email_badge = ""
    if new_emails:
        new_email_badge = (
            f"&nbsp;|&nbsp;<strong style='color:#8b0030'>{len(new_emails)} new "
            f"contact{'s' if len(new_emails) != 1 else ''}</strong>"
        )

    html = f"""
<html><body style="font-family:Arial,sans-serif;font-size:14px;color:#222">
<h2 style="color:#19473c">GrantConnect Opportunities — {run_date}</h2>
<p style="margin:8px 0">
  <strong>{len(results)}</strong> grant opportunities scraped &nbsp;|&nbsp;
  <strong>{len(with_emails)}</strong> with contact emails &nbsp;|&nbsp;
  <strong>{len(all_emails_set)}</strong> unique email addresses
  {new_go_badge}{new_email_badge}
</p>
{"<p style='background:#fff3f4;border-left:4px solid #ffadb5;padding:8px 12px;margin:8px 0'>"
 "<strong>★ New opportunities today:</strong> " + ", ".join(r["go_id"] for r in sorted(results, key=lambda x: x["close_date"]["sort"]) if r["url"] in new_go_urls) + "</p>"
 if new_go_urls else ""}
{"<p style='background:#fff3f4;border-left:4px solid #ffadb5;padding:8px 12px;margin:8px 0'>"
 "<strong>★ New contacts this run:</strong> " + ", ".join(sorted(new_emails)) + "</p>"
 if new_emails else ""}
<table style="border-collapse:collapse;width:100%">
<thead>
<tr style="background:#19473c;color:#fff">
  <th style="padding:6px 8px;text-align:left">GO ID</th>
  <th style="padding:6px 8px;text-align:left">Agency</th>
  <th style="padding:6px 8px;text-align:left">Title</th>
  <th style="padding:6px 8px;text-align:left">Closes</th>
  <th style="padding:6px 8px;text-align:left">Contact Email</th>
</tr>
</thead>
<tbody>
{rows_html}
</tbody>
</table>
<hr/>
<p style="color:#888;font-size:12px">
  ★ = new contact address not seen in previous runs &nbsp;|&nbsp;
  Registry file: grants_email_registry.xlsx &nbsp;|&nbsp;
  Generated {datetime.now(AEST).strftime("%Y-%m-%d %H:%M")}
</p>
</body></html>
"""

    text_lines = [f"GrantConnect Opportunities — {run_date}", ""]
    if new_emails:
        text_lines += ["NEW EMAILS THIS RUN:", *sorted(new_emails), ""]
    for r in sorted(results, key=lambda x: x["close_date"]["sort"]):
        marker = "NEW: " if any(e.lower() in new_emails for e in r["emails"]) else "     "
        emails_str = ", ".join(r["emails"]) if r["emails"] else "no email found"
        text_lines.append(f"{marker}{r['close_date']['display']} | {r['go_id']} | {r['agency']} | {emails_str}")
    plain = "\n".join(text_lines)

    new_count_str = f", {len(new_emails)} new" if new_emails else ""

    # Outer container: mixed (supports both body alternatives + attachment)
    msg = MIMEMultipart("mixed")
    msg["Subject"] = (
        f"GrantConnect Opportunities — {run_date} "
        f"({len(with_emails)} with contacts{new_count_str})"
    )
    msg["From"] = f"Market Analysis Tool <{CONFIG['from_email']}>"
    msg["To"] = CONFIG["to_email"]

    # Inner alternative block for plain/html body
    body = MIMEMultipart("alternative")
    body.attach(MIMEText(plain, "plain"))
    body.attach(MIMEText(html, "html"))
    msg.attach(body)

    # Attach the Excel registry
    registry_path = Path(CONFIG["registry_file"])
    if registry_path.exists():
        with open(registry_path, "rb") as f:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header("Content-Disposition", f'attachment; filename="{registry_path.name}"')
        msg.attach(part)

    log.info(f"Sending email to {CONFIG['to_email']}...")
    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
            smtp.login(CONFIG["from_email"], password)
            smtp.sendmail(CONFIG["from_email"], CONFIG["to_email"], msg.as_string())
    except smtplib.SMTPAuthenticationError:
        with smtplib.SMTP("smtp.gmail.com", 587) as smtp:
            smtp.ehlo()
            smtp.starttls()
            smtp.ehlo()
            smtp.login(CONFIG["from_email"], password)
            smtp.sendmail(CONFIG["from_email"], CONFIG["to_email"], msg.as_string())
    log.info("Email sent successfully.")


# ── Step 5: Generate HTML page ───────────────────────────────────────────────

def generate_html(results: list[dict], new_emails: set[str], new_go_urls: set[str]) -> str:
    """Build a standalone HTML page listing all current grant opportunities."""
    run_dt = datetime.now(AEST)
    run_date = run_dt.strftime("%d %B %Y")
    run_ts = run_dt.strftime("%Y-%m-%d %H:%M")
    all_emails_set = {e for r in results for e in r["emails"]}

    def esc(s): return html_mod.escape(str(s))

    # All opportunities sorted by close date
    rows = ""
    for r in sorted(results, key=lambda x: x["close_date"]["sort"]):
        is_new_go = r["url"] in new_go_urls
        is_new_email = any(e.lower() in new_emails for e in r["emails"])
        new_row = ' class="row-new"' if (is_new_go or is_new_email) else ""
        new_badge_html = ' <span class="new-badge">NEW</span>' if is_new_go else ""

        if r["emails"]:
            email_cells = []
            for email in r["emails"]:
                if email.lower() in new_emails:
                    email_cells.append(
                        f'<a href="mailto:{esc(email)}" class="email-new" title="New this run">★&nbsp;{esc(email)}</a>'
                    )
                else:
                    email_cells.append(f'<a href="mailto:{esc(email)}" class="email-link">{esc(email)}</a>')
            emails_html = "<br>".join(email_cells)
        else:
            emails_html = '<em style="color:#aaa">—</em>'

        rows += (
            f'<tr{new_row}>'
            f'<td><a href="{r["url"]}" target="_blank" rel="noopener">{esc(r["go_id"])}</a>{new_badge_html}</td>'
            f'<td>{esc(r["agency"])}</td>'
            f'<td>{esc(r["title"])}</td>'
            f'<td>{esc(r["category"])}</td>'
            f'<td data-sort="{r["close_date"]["sort"]}">{esc(r["close_date"]["display"])}</td>'
            f'<td>{emails_html}</td>'
            f'</tr>\n'
        )

    new_banner = ""
    if new_go_urls or new_emails:
        parts = []
        if new_go_urls:
            n = len(new_go_urls)
            parts.append(f'<strong>{n} new opportunit{"ies" if n != 1 else "y"} today</strong>')
        if new_emails:
            n = len(new_emails)
            email_list = ", ".join(f'<a href="mailto:{e}">{e}</a>' for e in sorted(new_emails))
            parts.append(f'<strong>{n} new contact{"s" if n != 1 else ""}:</strong> {email_list}')
        new_banner = f'<div class="banner-new">★ ' + " &nbsp;|&nbsp; ".join(parts) + "</div>"

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>GrantConnect Opportunities — {run_date} </title>
<style>
  :root {{
    --dark:       #19473c;
    --mid:        #1f735e;
    --pink:       #ffadb5;
    --pink-bg:    #fff3f4;
    --border:     #d0e0dc;
  }}
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: Arial, sans-serif; font-size: 14px; color: #222; background: #f0f5f4; }}
  header {{ background: var(--dark); color: #fff; padding: 18px 24px; }}
  header h1 {{ font-size: 20px; font-weight: bold; }}
  header p {{ font-size: 13px; color: #a8cec7; margin-top: 4px; }}
  header a {{ color: #a8cec7; }}
  .container {{ max-width: 1400px; margin: 0 auto; padding: 20px 16px; }}
  .stats {{ display: flex; gap: 16px; flex-wrap: wrap; margin-bottom: 16px; }}
  .stat {{ background: #fff; border: 1px solid var(--border); border-radius: 6px;
           padding: 10px 18px; text-align: center; }}
  .stat strong {{ display: block; font-size: 22px; color: var(--dark); }}
  .stat span {{ font-size: 12px; color: #888; }}
  .banner-new {{ background: var(--pink-bg); border-left: 4px solid var(--pink);
                 padding: 10px 14px; margin-bottom: 16px; border-radius: 0 6px 6px 0;
                 font-size: 13px; }}
  .search-bar {{ margin-bottom: 12px; }}
  .search-bar input {{ width: 100%; max-width: 400px; padding: 7px 12px;
                       border: 1px solid var(--border); border-radius: 4px; font-size: 14px; }}
  .search-bar input:focus {{ outline: none; border-color: var(--mid); box-shadow: 0 0 0 2px rgba(31,115,94,0.15); }}
  .table-wrap {{ overflow-x: auto; background: #fff; border-radius: 6px;
                 border: 1px solid var(--border); }}
  table {{ width: 100%; border-collapse: collapse; }}
  thead tr {{ background: var(--dark); color: #fff; }}
  th {{ padding: 9px 10px; text-align: left; font-size: 13px; cursor: pointer; white-space: nowrap; }}
  th:hover {{ background: var(--mid); }}
  td {{ padding: 7px 10px; border-bottom: 1px solid #e8f0ee; vertical-align: top; }}
  td:nth-child(5) {{ white-space: nowrap; }}
  tr:last-child td {{ border-bottom: none; }}
  tr.row-new {{ background: var(--pink-bg); }}
  tr:hover td {{ background: #eef5f3; }}
  tr.row-new:hover td {{ background: #ffe8ea; }}
  a {{ color: var(--mid); text-decoration: none; }}
  a:hover {{ text-decoration: underline; }}
  .email-new {{ color: #c0004a; font-weight: bold; }}
  .footer {{ text-align: center; color: #aaa; font-size: 12px; margin-top: 24px; padding-bottom: 24px; }}
  .hidden {{ display: none; }}
  th .sort-icon {{ font-size: 10px; margin-left: 4px; opacity: 0.6; }}
  .new-badge {{ background: var(--pink); color: #8b0030; font-size: 11px; font-weight: bold;
                padding: 1px 6px; border-radius: 3px; margin-left: 6px; vertical-align: middle; }}
</style>
</head>
<body>
<header>
  <h1>GrantConnect Opportunities — {run_date}</h1>
  <p>Scraped from <a href="https://www.grants.gov.au/Go/List" style="color:#7ab" target="_blank">grants.gov.au</a>
     &nbsp;|&nbsp; Updated {run_ts} {datetime.now(AEST).strftime("%Z")} &nbsp;|&nbsp;
     <a href="{CONFIG['pages_url']}" style="color:#7ab">Permalink</a></p>
</header>

<div class="container">
  <div class="stats">
    <div class="stat"><strong>{len(results)}</strong><span>opportunities</span></div>
    <div class="stat"><strong>{sum(1 for r in results if r["emails"])}</strong><span>with contact email</span></div>
    <div class="stat"><strong>{len(all_emails_set)}</strong><span>unique addresses</span></div>
    <div class="stat"><strong>{len(new_go_urls) if new_go_urls else "—"}</strong><span>new opportunities</span></div>
    <div class="stat"><strong>{len(new_emails) if new_emails else "—"}</strong><span>new contacts</span></div>
  </div>

  {new_banner}

  <div class="search-bar">
    <input type="text" id="searchInput" placeholder="Filter by agency, title, category, email…" oninput="filterTable()">
  </div>

  <div class="table-wrap">
    <table id="mainTable">
      <thead>
        <tr>
          <th onclick="sortTable(0)">GO ID <span class="sort-icon">⇅</span></th>
          <th onclick="sortTable(1)">Agency <span class="sort-icon">⇅</span></th>
          <th onclick="sortTable(2)">Title <span class="sort-icon">⇅</span></th>
          <th onclick="sortTable(3)">Category <span class="sort-icon">⇅</span></th>
          <th onclick="sortTable(4)">Closes <span class="sort-icon">⇅</span></th>
          <th onclick="sortTable(5)">Contact Email <span class="sort-icon">⇅</span></th>
        </tr>
      </thead>
      <tbody id="tableBody">
        {rows}
      </tbody>
    </table>
  </div>

  <div class="footer">
    ★ = new contact not seen in previous run &nbsp;|&nbsp;
    Data sourced from <a href="https://www.grants.gov.au">grants.gov.au</a> &nbsp;|&nbsp;
    Generated {run_ts}
  </div>
</div>

<script>
function filterTable() {{
  const q = document.getElementById('searchInput').value.toLowerCase();
  const rows = document.querySelectorAll('#tableBody tr');
  rows.forEach(row => {{
    row.classList.toggle('hidden', q && !row.textContent.toLowerCase().includes(q));
  }});
}}

let sortDir = {{}};
function sortTable(col) {{
  const tbody = document.getElementById('tableBody');
  const rows = Array.from(tbody.querySelectorAll('tr'));
  const asc = sortDir[col] !== true;
  sortDir = {{}};
  sortDir[col] = asc;
  // Update header arrow indicators
  document.querySelectorAll('th .sort-icon').forEach((el, i) => {{
    el.textContent = i === col ? (asc ? '↑' : '↓') : '⇅';
  }});
  rows.sort((a, b) => {{
    const ac = a.cells[col];
    const bc = b.cells[col];
    const av = ac?.dataset.sort ?? ac?.textContent.trim() ?? '';
    const bv = bc?.dataset.sort ?? bc?.textContent.trim() ?? '';
    return asc ? av.localeCompare(bv) : bv.localeCompare(av);
  }});
  rows.forEach(r => tbody.appendChild(r));
}}
// Default sort: closes soonest first (col 4)
window.addEventListener('DOMContentLoaded', () => sortTable(4));
</script>
</body>
</html>
"""


# ── Step 6: Write HTML locally (git workflow commits it to GitHub Pages) ───────

def save_html(html: str) -> None:
    path = Path(CONFIG["html_file"])
    path.write_text(html, encoding="utf-8")
    log.info(f"HTML written to {path}")


# ── Main ──────────────────────────────────────────────────────────────────────

async def main():
    import os
    password = CONFIG["gmail_app_password"] or os.environ.get("GMAIL_APP_PASSWORD", "")
    if not password:
        log.error("No Gmail App Password found. Set GMAIL_APP_PASSWORD env var or edit CONFIG.")
        sys.exit(1)

    log.info("=" * 60)
    log.info(f"GrantConnect scraper starting — {datetime.now(AEST).strftime('%Y-%m-%d %H:%M')}")

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        context = await browser.new_context(
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            )
        )

        opps = await collect_all_go_opportunities(context)
        urls = [o["url"] for o in opps]

        # Detect new grant opportunities vs previous run
        known_urls = load_known_go_urls()
        new_go_urls = set(urls) - known_urls
        log.info(f"New grant opportunities this run: {len(new_go_urls)}")

        log.info(f"Scraping {len(opps)} detail pages (concurrency={CONFIG['concurrency']})...")
        results = await scrape_all_details(context, opps)

        await browser.close()

    # Persist current URL set for next run's comparison
    save_known_go_urls(urls)

    with_emails = [r for r in results if r["emails"]]
    log.info(f"Scrape complete: {len(with_emails)}/{len(results)} opportunities have emails")

    new_emails, _ = update_registry(results)
    log.info(f"New emails this run: {len(new_emails)}")

    html_page = generate_html(results, new_emails, new_go_urls)
    save_html(html_page)

    send_email(results, new_emails, new_go_urls, password)
    log.info("Done.")


if __name__ == "__main__":
    asyncio.run(main())
